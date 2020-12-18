from openpyxl import load_workbook
workbook = load_workbook(filename="C:/Users/44753/Downloads/2018-2019 NBA Player Stats  NBAstuffer.xlsx")
sheet = workbook.active
import random       
row = random.randint(2,213)
def shooter(row):
    row = random.randint(1,213)
    global name
    name = sheet.cell(row,1).value
    global team
    team = sheet.cell(row,2).value
    global usp
    usp = sheet.cell(row,8).value
    global threep
    threep = sheet.cell(row,9).value
    global truep
    truep = sheet.cell(row,10).value
    global lv
    lv = (int(threep) * int(truep) * int(usp))
class Shooter:
    def __init__(self, name, team, usp, threep, truep,lv):
        self.name = name
        self.team = team
        self.usp = usp
        self.threep = threep
        self.truep = truep
        self.lv = lv
        
#player1 team
shooter(row)
s1 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s3 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s5 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s7 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s9 = Shooter(name, team, usp, threep, truep,lv)


#player2team
shooter(row)
s2 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s4 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s6 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s8 = Shooter(name, team, usp, threep, truep,lv)
shooter(row)
s10 = Shooter(name, team, usp, threep, truep,lv)

from tkinter import*
root = Tk()
root.geometry('405x750')
root.title("NBA SHOOT-OFF")
label_0 = Label(root, text = "SHOOT-OFF",width = 20,font =("bold", 20))
label_0.place(x=500,y=1)

#p1name
label_1 = Label(root, text = "Player1Name", width=20,font=("bold",10))
label_1.place(x=0,y=60) 
p1name = Entry(root)
p1_name = p1name.get()
p1name.place(x=240,y=60)

#p2name
label_2 = Label(root, text = "Player2Name", width=20,font=("bold",10))
label_2.place(x=0,y=90)
p2name = Entry(root)
p2_name = p2name.get()
p2name.place(x=240,y= 90)

w1 = 0
w2 = 0
winner = ''
rounds = 5


def start(p1_name,p2_name,rounds,w1,w2,winner):
    label_5 = Label(root, text = "Team " + str(p1name.get()), width = 20, font=("bold", 10))
    label_5.place(x=50,y = 170)
    label_8 = Label(root, text = "Team " + str(p2name.get()), width = 20, font=("bold", 10))
    label_8.place(x=50,y = 210)

    team1 = [s1.name, s3.name, s5.name, s7.name, s9.name]
    team2 = [s2.name, s4.name, s6.name, s8.name, s10.name]
    label_6 = Label(root, text = str(team1),font = ("bold",10))
    label_6.place(x=0,y=190)
    label_7 = Label(root,text = str(team2), font = ("bold",10))
    label_7.place(x=0,y=230)

    team1 = [s1, s3, s5, s7, s9]
    team2 = [s2, s4, s6, s8, s10]
    for i in range(0,5):
        label_4 = Label(root, text = str(rounds),font = ("bold",15))
        label_4.place(x=1300,y=60) 
        label_9 = Label(root, text = team1[i].name + " vs " + team2[i].name, font=("bold",10))
        label_9.place(x=800,y=155)
        if team1[i].lv > team2[i].lv:
                winner = team1[i].name
                w1 += 1
        else:
            winner = team2[i].name
            w2 += 1
            label_10 = Label(root, text = winner + " won", font =("bold",10))
            label_10.place(x=800,y=255)
        rounds-=1               
    if w1==w2:
        print("Game ended in draw", font = ("bold", 20))
    else:
        if w1>w2:
            victor = p1name.get()
        else:
            victor = p2name.get()
            print("Winner is" + victor, font = ("bold", 20))
            
bstart = Button(root,text="start", command = lambda: start(p1_name,p2_name,rounds,w1,w2,winner))
bstart.place(x=40,y=130)
label_3 = Label(root, text = "Round", width = 10, font = ("bold",15))
label_3.place(x=1215,y=60)
